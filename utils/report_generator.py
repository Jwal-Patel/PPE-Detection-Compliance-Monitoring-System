"""
Report generation for PPE compliance analytics.
Generates PDF, Excel, and JSON reports.
PHASE 4: Multi-format report generation.
"""

from typing import Dict, Optional
import pandas as pd
from datetime import datetime
import json
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Generate reports in multiple formats.
    """
    
    @staticmethod
    def generate_csv_report(
        df: pd.DataFrame,
        filename: str
    ) -> bytes:
        """
        Generate CSV report.
        
        Args:
            df: DataFrame with data
            filename: Output filename
            
        Returns:
            CSV data as bytes
        """
        try:
            csv_data = df.to_csv(index=False)
            logger.info(f"CSV report generated: {filename}")
            return csv_data.encode('utf-8')
        except Exception as e:
            logger.error(f"CSV generation error: {str(e)}")
            return b""
    
    @staticmethod
    def generate_json_report(
        data: Dict,
        filename: str
    ) -> bytes:
        """
        Generate JSON report.
        
        Args:
            data: Data dictionary
            filename: Output filename
            
        Returns:
            JSON data as bytes
        """
        try:
            json_data = json.dumps(data, indent=2, default=str)
            logger.info(f"JSON report generated: {filename}")
            return json_data.encode('utf-8')
        except Exception as e:
            logger.error(f"JSON generation error: {str(e)}")
            return b""
    
    @staticmethod
    def generate_excel_report(
        df: pd.DataFrame,
        filename: str,
        sheet_name: str = "Report"
    ) -> bytes:
        """
        Generate Excel report.
        
        Args:
            df: DataFrame with data
            filename: Output filename
            sheet_name: Excel sheet name
            
        Returns:
            Excel data as bytes
        """
        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Write headers with formatting
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Format header row
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"Excel report generated: {filename}")
            return output.getvalue()
        
        except Exception as e:
            logger.error(f"Excel generation error: {str(e)}")
            return b""
    
    @staticmethod
    def generate_pdf_report(
        df: pd.DataFrame,
        title: str,
        filename: str,
        stats: Optional[Dict] = None
    ) -> bytes:
        """
        Generate PDF report with charts and statistics.
        
        Args:
            df: DataFrame with data
            title: Report title
            filename: Output filename
            stats: Optional statistics dictionary
            
        Returns:
            PDF data as bytes
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib import colors
            from datetime import datetime
            import io
            
            # Create PDF
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#667eea'),
                spaceAfter=30,
                alignment=1  # Center
            )
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Report date
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                alignment=1
            )
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}", date_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Statistics section if provided
            if stats:
                elements.append(Paragraph("Summary Statistics", styles['Heading2']))
                
                stats_data = [
                    ["Metric", "Value"],
                    ["Total Detections", str(stats.get('total_detections', 0))],
                    ["Total Workers", str(stats.get('total_workers', 0))],
                    ["Compliant", str(stats.get('compliant', 0))],
                    ["Partial", str(stats.get('partial', 0))],
                    ["Non-Compliant", str(stats.get('non_compliant', 0))],
                    ["Compliance Rate", f"{stats.get('compliance_rate', 0):.1f}%"],
                ]
                
                stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                ]))
                
                elements.append(stats_table)
                elements.append(Spacer(1, 0.3*inch))
            
            # Data table
            elements.append(Paragraph("Detailed Data", styles['Heading2']))
            elements.append(Spacer(1, 0.2*inch))
            
            # Convert DataFrame to table
            data = [list(df.columns)] + df.values.tolist()
            
            table = Table(data, colWidths=[1*inch] * len(df.columns))
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            
            logger.info(f"PDF report generated: {filename}")
            return buffer.getvalue()
        
        except ImportError:
            logger.warning("reportlab not installed. Install with: pip install reportlab")
            return b""
        
        except Exception as e:
            logger.error(f"PDF generation error: {str(e)}")
            return b""